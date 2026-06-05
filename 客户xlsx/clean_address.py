import openpyxl
import re
import shutil
import os

file_path = r'D:\code空间\待确定直播与否.xlsx'
backup_path = r'D:\code空间\待确定直播与否_backup.xlsx'

shutil.copy2(file_path, backup_path)
print(f"备份已创建: {backup_path}")

wb = openpyxl.load_workbook(file_path)
ws = wb.active

updated_count = 0
keep_count = 0

for row in range(2, ws.max_row + 1):
    b_val = ws.cell(row=row, column=2).value
    if b_val is None:
        continue

    text = str(b_val)

    if '### 1.' in text:
        m = re.search(r'###\s*1\.\s*地址[：:]?\s*(.*?)(?:###\s*2\.|\Z)', text, re.DOTALL)
        if m:
            clean_addr = m.group(1).strip()
            clean_addr = re.sub(r'【[^】]*】', '', clean_addr).strip()
            ws.cell(row=row, column=2).value = clean_addr
            updated_count += 1
            print(f"第{row}行: 已清理 -> {clean_addr}")
    else:
        keep_count += 1

wb.save(file_path)
print(f"\n处理完成！共处理 {ws.max_row - 1} 行数据")
print(f"已清理: {updated_count} 行")
print(f"保持原样: {keep_count} 行")