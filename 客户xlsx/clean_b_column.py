import re
import openpyxl
import shutil
from datetime import datetime

backup_path = r'd:\code空间\客户xlsx\汇总客户_备份.xlsx'
filepath = r'd:\code空间\客户xlsx\汇总客户.xlsx'

# ── 第1步：备份 ──
ts = datetime.now().strftime('%Y%m%d_%H%M%S')
auto_backup = rf'd:\code空间\客户xlsx\_自动备份_{ts}_汇总客户.xlsx'
try:
    shutil.copy2(filepath, auto_backup)
    print(f'✅ 已备份当前文件 → {auto_backup}')
except FileNotFoundError:
    print('⚠️  当前汇总客户.xlsx不存在，跳过备份')

# ── 第2步：恢复原始数据 ──
shutil.copy2(backup_path, filepath)
print(f'✅ 已从备份恢复原始数据 → {filepath}')

# ── 第3步：开始清洗 ──
wb = openpyxl.load_workbook(filepath)
ws = wb.active

JUNK_TRIGGERS = (
    r'曾发布|未发布|发布过|招淘宝天猫|商标归属|岗位招聘|招聘链接|招聘主播|'
    r'无杭州|无匹配|无关联|无对应|无岗位|营业执照|'
    r'仅招聘|招聘熨|招聘电商|商标注册|实际运营|疑似挂靠|疑似'
)

def strip_brackets(s):
    s = re.sub(r'【[^】]*】', '', s)
    s = re.sub(r'^\d+\.(?:地址|招淘宝天猫运营的地址)[：:]?\s*', '', s)
    s = re.sub(r'\s+', ' ', s)
    s = s.strip().rstrip('，,;； ')
    # 统一前缀：去掉开头的"浙江省"
    if s.startswith('浙江省'):
        s = s[3:]
    return s

def extract_address(text):
    if not text:
        return text
    text = text.strip()
    if not text:
        return text

    # ── 情况1: "1.地址：{地址} 2.boss/58..." ──
    m = re.search(r'^1\.(?:地址|招淘宝天猫运营的地址)[：:]?\s*(.*?)(?=\s+2\.)', text)
    if m:
        addr = m.group(1).strip().rstrip('，, ')
        if addr:
            # 如果地址末尾含（BOSS直聘/未发布/曾发布），截断
            m2 = re.match(r'^([^（]*?)（(?:BOSS直聘|曾发布|未发布)', addr)
            if m2:
                addr = m2.group(1).strip().rstrip('，, ')
            if addr:
                # 如果提取到的地址包含否定关键词，说明没有真实地址，置空
                if re.search(r'无杭州|无岗位|无地址|无办公|无关联', addr):
                    return ''
                addr = strip_brackets(addr)
                if addr:
                    return addr

    # ── 情况2: "1.地址：{地址}，该地址为..." ──
    m = re.search(r'^1\.地址[：:]?\s*(.*?)(?:[，,]\s*(?:该地址|未发布|曾发布)|$)', text)
    if m:
        addr = m.group(1).strip().rstrip('，, ')
        if addr:
            addr = strip_brackets(addr)
            if addr:
                return addr

    # ── 情况3: "1.招淘宝天猫运营的地址：{地址}【..." ──
    m = re.search(r'^1\.招淘宝天猫运营的地址[：:]?\s*(.*?)(?:【|，|,|$)', text)
    if m:
        addr = m.group(1).strip().rstrip('，, ')
        if addr:
            addr = strip_brackets(addr)
            if addr:
                return addr

    # ── 情况4: 纯地址 ──
    junk_pattern = re.compile(
        r'(曾发布|未发布|发布过|招淘宝天猫|商标归属|'
        r'岗位招聘|招聘链接|https?://|'
        r'无杭州|无匹配|无关联|无对应|无岗位|'
        r'营业执照|招聘主播|招聘电商|招聘熨|仅招聘|'
        r'商标注册|实际运营主体|关联公司|BOSS直聘|'
        r'zhipin|58同城)'
    )
    is_pure = not junk_pattern.search(text)
    if (text.startswith('杭州市') or text.startswith('浙江省杭州市')) and is_pure:
        return strip_brackets(text)

    # ── 情况5: 以"杭州市"开头但包含冗余信息 ──
    for prefix in ['浙江省杭州市', '杭州市']:
        if text.startswith(prefix):
            addr_body = r'[^，,。]*?(?:【[^】]*】)*(?:（[^）]*）)*(?:\([^)]*\))*[^，,。]*?'

            # 5a: 逗号 + 触发词（保留合法括号，过滤脏结果）
            m = re.search(rf'^({prefix}{addr_body})\s*[，,]\s*(?:{JUNK_TRIGGERS})', text)
            if m:
                addr = m.group(1).strip()
                if addr and not junk_pattern.search(addr):
                    addr = strip_brackets(addr)
                    if addr:
                        return addr

            # 5e: （曾发布/未发布...）在地址末尾（优先处理括号内脏词）
            m = re.search(rf'^({prefix}[^（]*?)（(?:曾发布|未发布)', text)
            if m:
                addr = m.group(1).strip()
                if addr and not junk_pattern.search(addr):
                    addr = strip_brackets(addr)
                    if addr:
                        return addr

            # 5b: 空格 + 触发词（过滤脏结果）
            m = re.search(rf'^({prefix}{addr_body})\s+(?:{JUNK_TRIGGERS})', text)
            if m:
                addr = m.group(1).strip()
                if addr and not junk_pattern.search(addr):
                    addr = strip_brackets(addr)
                    if addr:
                        return addr

            # 5c: 地址后直接接"未发布/曾发布"（排除（和【语境）
            m = re.search(rf'^({prefix}[^（【]*?)(?:未发布|曾发布)', text)
            if m:
                addr = m.group(1).strip()
                if addr:
                    addr = strip_brackets(addr)
                    if addr:
                        return addr

            # 5d: https://...
            m = re.search(rf'^({prefix}{addr_body})\s*https?://', text)
            if m:
                addr = m.group(1).strip().rstrip('，, ')
                if addr:
                    addr = strip_brackets(addr)
                    if addr:
                        return addr

            # 5f: 用关键词截断
            cut_keywords = [
                ' 曾发布', ' 未发布', ' 发布过',
                ' 无杭州', ' 无关联', ' 无对应', ' 无岗位',
                ' 营业执照', ' 招聘主播', ' 招聘电商',
                ' BOSS直聘', ' zhipin', ' 58同城',
            ]
            for kw in cut_keywords:
                idx = text.find(kw)
                if idx > len(prefix):
                    addr = text[:idx].strip().rstrip('，, ')
                    if addr and not junk_pattern.search(addr):
                        addr = strip_brackets(addr)
                        if addr:
                            return addr

            # 5h: "杭州市...【...】未发布/曾发布..."
            m = re.search(rf'^({prefix}[^【]*?(?:【[^】]*】)*)\s*(?:未发布|曾发布)', text)
            if m:
                addr = m.group(1).strip()
                if addr and not junk_pattern.search(addr):
                    addr = strip_brackets(addr)
                    if addr:
                        return addr

            # 5i: BOSS直聘等关键词在括号内
            m = re.search(rf'^({prefix}[^（]*?)（(?:BOSS直聘)', text)
            if m:
                addr = m.group(1).strip()
                if addr and not junk_pattern.search(addr):
                    addr = strip_brackets(addr)
                    if addr:
                        return addr

            # 5j: 【】注释内部包含脏词（未发布/曾发布），提取到第一个【】之前
            m = re.search(rf'^({prefix}[^【]+?)【', text)
            if m:
                addr = m.group(1).strip()
                if addr and not junk_pattern.search(addr):
                    addr = strip_brackets(addr)
                    if addr:
                        return addr

    # ── 情况6: 包含"地址：" ──
    m = re.search(r'地址[：:]?\s*(杭州市[^，,。]*?(?:室|楼|号|层|幢|座|区|园|街|巷))', text)
    if m:
        addr = m.group(1).strip()
        if not junk_pattern.search(addr):
            addr = strip_brackets(addr)
            if addr:
                return addr

    # ── 情况7: 纯净内容 ──
    if is_pure:
        return strip_brackets(text)

    # ── 情况8: 兜底 ──
    return strip_brackets(text)


changes = []
for row in ws.iter_rows(min_row=1, max_row=ws.max_row, min_col=2, max_col=2):
    cell = row[0]
    old_val = cell.value
    new_val = extract_address(old_val)
    if old_val != new_val:
        changes.append((cell.row, old_val, new_val))
        cell.value = new_val

wb.save(filepath)

print(f'共处理 {ws.max_row} 行')
print(f'共修改 {len(changes)} 行')
print()
print('=== 修改示例（前20行）===')
for i, (row_num, old, new) in enumerate(changes[:20]):
    old_short = str(old)[:90] + '...' if len(str(old)) > 90 else str(old)
    new_short = str(new)[:90] + '...' if len(str(new)) > 90 else str(new)
    print(f'\n第{row_num}行:')
    print(f'  [前] {old_short}')
    print(f'  [后] {new_short}')

print()
print('=== 验证：检查是否还有脏数据 ===')
junk_kw = ['曾发布','未发布','招淘宝天猫','商标归属','岗位招聘','https://','http://','无杭州','招聘主播','无关联','无对应','BOSS直聘','zhipin','58同城']
remaining_dirty = 0
for row_obj in ws.iter_rows(min_row=1, max_row=ws.max_row, min_col=2, max_col=2, values_only=True):
    b = row_obj[0]
    if b and any(kw in b for kw in junk_kw):
        remaining_dirty += 1
        if remaining_dirty <= 20:
            print(f'  [{b[:120]}]')
print(f'剩余脏行数: {remaining_dirty}')