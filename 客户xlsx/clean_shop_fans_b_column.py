import re
import shutil
import openpyxl
import os
from datetime import datetime

file_path = r'd:\code空间\客户xlsx\店铺粉丝量完整清单-2026.xlsx'
temp_path = r'd:\code空间\客户xlsx\_清洗中_店铺粉丝量完整清单-2026.xlsx'

# 备份
ts = datetime.now().strftime('%Y%m%d_%H%M%S')
backup_path = rf'd:\code空间\客户xlsx\_自动备份_{ts}_店铺粉丝量完整清单.xlsx'
try:
    shutil.copy2(file_path, backup_path)
    print(f'✅ 已备份 → {backup_path}')
except PermissionError:
    print('⚠️  无法备份（文件被占用），但仍然尝试读取清洗')

wb = openpyxl.load_workbook(file_path)
ws = wb.active

JUNK_TRIGGERS = (
    r'曾发布|未发布|发布过|发布主播|该地址为|非岗位填报|'
    r'招淘宝天猫|商标归属|商标注册|岗位招聘|招聘链接|招聘主播|'
    r'招聘电商|招聘运营|招聘直播|招聘熨|仅招聘|'
    r'无杭州|无匹配|无关联|无对应|无岗位|无商标|'
    r'营业执照|实际运营|疑似挂靠|疑似|'
    r'有效商标|有效招聘|有效岗位|'
    r'无有效|无BOSS|无杭州地区|'
    r'未检索到|未查询到|未查询|暂未查询|'
    r'无关联公司|无岗位招聘|商家未提供|'
    r'自然人|个体工商户|'
    r'https?://|BOSS直聘|zhipin|58同城'
)

def clean_address(text):
    if not text:
        return text
    text = str(text).strip()
    if not text:
        return text

    # 去掉开头的数字序号，如 "1.地址："、"1.招淘宝天猫运营的地址："
    text = re.sub(r'^\d+\.(?:地址|招淘宝天猫运营的地址)[：:]?\s*', '', text)

    # 去掉末尾多余的空格和标点
    text = text.strip().rstrip('，,;； 　')

    # 替换全角空格
    text = text.replace('\u3000', ' ')

    # 找地址开头：浙江省杭州市 或 杭州市
    addr_match = re.search(r'(浙江省杭州市|杭州市)', text)
    if not addr_match:
        return text

    start = addr_match.start()
    addr = text[start:]

    # 按各种触发词截断
    cut_patterns = [
        # 逗号 + 触发词
        rf'[，,]\s*(?:{JUNK_TRIGGERS})',
        # 空格 + 触发词
        rf'\s+(?:{JUNK_TRIGGERS})',
        # （曾发布 / 未发布
        rf'（(?:曾发布|未发布|发布过)',
        # 直接接未发布/曾发布（无分隔符）
        r'(未发布|曾发布|发布过)',
        # 【括号
        r'【',
        # 换行后的内容
        r'\n',
    ]

    best_pos = len(addr)
    for pat in cut_patterns:
        m = re.search(pat, addr)
        if m and m.start() < best_pos:
            best_pos = m.start()

    addr = addr[:best_pos].strip().rstrip('，,;； \t')

    # 去掉地址中残留的【...】
    addr = re.sub(r'【[^】]*】', '', addr).strip()

    # 去掉地址中残留的（...）如果括号内包含脏词
    addr = re.sub(r'（(?:曾发布|未发布|发布过)[^）]*）', '', addr).strip()
    addr = re.sub(r'（[^）]*?(?:曾发布|未发布|发布过)[^）]*）', '', addr).strip()

    # 再次清理尾部
    addr = addr.strip().rstrip('，,;； \t')

    # 如果清理后以"无"开头（如"无杭州"），说明没有真实地址
    if re.match(r'^(无|曾发布|未发布)', addr):
        return ''

    return addr

changes = []
for row in ws.iter_rows(min_row=1, max_row=ws.max_row, min_col=2, max_col=2):
    cell = row[0]
    old_val = cell.value
    new_val = clean_address(old_val)
    if old_val != new_val:
        changes.append((cell.row, old_val, new_val))
        cell.value = new_val

wb.save(temp_path)
wb.close()

# 尝试用临时文件替换原文件
try:
    shutil.move(temp_path, file_path)
    print(f'✅ 已保存并替换原文件')
except PermissionError:
    print(f'⚠️  原文件被占用，清洗结果已保存到: {temp_path}')
    print(f'   请关闭 Excel 后手动将 {temp_path} 重命名为原文件名')

print(f'\n共处理 {ws.max_row} 行')
print(f'共修改 {len(changes)} 行')

print('\n=== 修改示例（前30行）===')
for i, (row_num, old, new) in enumerate(changes[:30]):
    if new:
        print(f'\n第{row_num}行:')
        old_short = str(old)[:80] + '...' if len(str(old)) > 80 else str(old)
        print(f'  [前] {old_short}')
        print(f'  [后] {new}')
    else:
        print(f'\n第{row_num}行:')
        old_short = str(old)[:80] + '...' if len(str(old)) > 80 else str(old)
        print(f'  [前] {old_short}')
        print(f'  [后] (已置空 - 无有效地址)')

# 验证
print('\n=== 验证：检查是否还有脏数据 ===')
junk_kw = ['曾发布','未发布','招淘宝天猫','商标归属','岗位招聘','https://','http://',
           '无杭州','无匹配','无关联','无对应','无岗位','招聘主播','无BOSS',
           'BOSS直聘','zhipin','58同城','疑似挂靠','该地址为']
remaining_dirty = 0
for row_obj in ws.iter_rows(min_row=1, max_row=ws.max_row, min_col=2, max_col=2, values_only=True):
    b = row_obj[0]
    if b and any(kw in b for kw in junk_kw):
        remaining_dirty += 1
        if remaining_dirty <= 10:
            print(f'  [{str(b)[:120]}]')
print(f'剩余脏行数: {remaining_dirty}')