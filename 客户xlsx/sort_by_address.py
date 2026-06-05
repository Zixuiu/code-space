import openpyxl
import shutil

file_path = r'd:\code空间\新建文件夹 (2)\汇总客户_清爽版_A去重_B排序.xlsx'
output_path = r'd:\code空间\新建文件夹 (2)\汇总客户_清爽版_A去重_B排序_已排序.xlsx'

shutil.copy2(file_path, output_path)
print(f"复制文件到: {output_path}")

wb = openpyxl.load_workbook(output_path)
ws = wb.active

rows_with_address = []
rows_without_address = []

header = [ws.cell(row=1, column=j).value for j in range(1, ws.max_column + 1)]

for row in range(2, ws.max_row + 1):
    row_data = []
    has_address = False
    
    for col in range(1, ws.max_column + 1):
        cell_value = ws.cell(row=row, column=col).value
        row_data.append(cell_value)
        
        if col == 2 and cell_value:
            addr_str = str(cell_value).strip()
            if addr_str and '未找到' not in addr_str and addr_str.lower() != 'none':
                has_address = True
    
    if has_address:
        rows_with_address.append(row_data)
    else:
        rows_without_address.append(row_data)

ws.delete_rows(2, ws.max_row)

current_row = 2
for row_data in rows_with_address:
    for col, value in enumerate(row_data, 1):
        ws.cell(row=current_row, column=col).value = value
    current_row += 1

for row_data in rows_without_address:
    for col, value in enumerate(row_data, 1):
        ws.cell(row=current_row, column=col).value = value
    current_row += 1

wb.save(output_path)
print(f"\n处理完成！")
print(f"有地址的行数: {len(rows_with_address)}")
print(f"未找到地址的行数: {len(rows_without_address)}")
print(f"结果已保存到: {output_path}")