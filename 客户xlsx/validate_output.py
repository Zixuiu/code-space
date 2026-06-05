from openpyxl import load_workbook
wb = load_workbook(r"d:\code空间\客户xlsx\店铺粉丝量完整清单-2026.xlsx")
ws = wb.active
print(f"总行数（含表头）: {ws.max_row}")
print()

long_fans = 0
empty_fans = 0
print("=== 粉丝值异常样例 ===")
for i, row in enumerate(ws.iter_rows(min_row=2, values_only=True), 2):
    fans = str(row[1]) if row[1] else ""
    if len(fans) > 20:
        long_fans += 1
        if long_fans <= 15:
            print(f"  行{i}: {row}")
    if not fans.strip():
        empty_fans += 1

print()
print(f"粉丝值过长的: {long_fans} 条")
print(f"粉丝值为空的: {empty_fans} 条")
print(f"有效数据: {ws.max_row - 1 - empty_fans} 条")
print()
print("文件: d:\\code空间\\客户xlsx\\店铺粉丝量完整清单-2026.xlsx")
wb.close()