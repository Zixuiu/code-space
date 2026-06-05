import openpyxl

input_path = r'd:\code空间\客户xlsx\店铺粉丝量完整清单-2026.xlsx'
output_path = r'd:\code空间\客户xlsx\店铺粉丝量完整清单-2026_清洗结果.xlsx'

wb = openpyxl.load_workbook(input_path)
ws = wb['店铺粉丝量清单-已确定位置']

seen = set()
addresses = []

for row in ws.iter_rows(min_row=1, values_only=True):
    addr = str(row[1]).strip() if row[1] else ''

    if not addr:
        continue

    if addr in seen:
        continue
    seen.add(addr)

    addresses.append([addr])

wb_out = openpyxl.Workbook()
ws_out = wb_out.active
ws_out.title = '地址清单'
ws_out.append(['地址/备注'])

for a in addresses:
    ws_out.append(a)

ws_out.column_dimensions['A'].width = 100

wb_out.save(output_path)

print(f'B列有地址的行(原始): 680')
print(f'去重后保留: {len(addresses)}')
print(f'已保存: {output_path}')
print()
print('--- 前10个地址 ---')
for i, a in enumerate(addresses[:10]):
    print(f'{i+1}. {a[0][:60]}')