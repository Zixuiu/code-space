import openpyxl

wb = openpyxl.load_workbook(r'd:\code空间\客户xlsx\汇总客户.xlsx')
ws = wb.active

junk_kw = ['曾发布','未发布','招淘宝天猫','商标归属','岗位招聘','https://','http://','无杭州','招聘主播','无关联','无对应','BOSS直聘','zhipin','58同城']

print(f'总行数: {ws.max_row}')
print()

dirty_rows = []
for row in ws.iter_rows(min_row=1, max_row=ws.max_row, min_col=2, max_col=2):
    cell = row[0]
    b = cell.value
    if b and any(kw in b for kw in junk_kw):
        dirty_rows.append((cell.row, b))

print(f'剩余脏行数: {len(dirty_rows)}')
print()
for row_num, val in dirty_rows:
    print(f'--- 第{row_num}行 ---')
    print(f'  {val}')
    print()