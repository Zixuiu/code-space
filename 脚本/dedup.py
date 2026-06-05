from openpyxl import load_workbook

wb = load_workbook(r'd:\code空间\总路线.xlsx')
ws = wb.active

seen = set()
rows_to_delete = []

for row in range(2, ws.max_row + 1):
    name = ws.cell(row=row, column=1).value
    if name in seen:
        rows_to_delete.append(row)
    else:
        seen.add(name)

for r in reversed(rows_to_delete):
    ws.delete_rows(r)

output_path = r'd:\code空间\总路线_去重.xlsx'
wb.save(output_path)

print(f"原记录数: {ws.max_row + len(rows_to_delete) - 1}")
print(f"去重后记录数: {ws.max_row - 1}")
print(f"删除重复: {len(rows_to_delete)} 条")